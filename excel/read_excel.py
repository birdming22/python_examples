#!/usr/bin/python
# -*- coding: UTF-8 -*-

from datetime import datetime, timedelta
import openpyxl
import random


def generate_weekdays(start_date):
    while True:
        if start_date.weekday() == 0: # Monday
            yield start_date
            start_date += timedelta(days=2) # Wednesday
        elif start_date.weekday() == 2: # Wednesday
            yield start_date
            start_date += timedelta(days=5) # Next Monday
        else:
            start_date += timedelta(days=1)


def main():
    # 讀取 Excel 檔案
    excel_file = openpyxl.load_workbook('template.xlsx')

    # 選擇工作表
    sheet = excel_file['工作表1']


    # 設定時間格式
    date_format = '%Y/%m/%d %H:%M:%S'


    template = []
    for row in sheet.iter_rows(4, values_only=True):
        template.append(row)

    # 設定日期為 2022 年 1 月 1 日
    start_date = datetime(2021, 6, 1)
    row_num = 0
    row_offset = 4
    for working_date in generate_weekdays(start_date):
        year = int(working_date.strftime('%Y'))
        if year >= 2022:
            break
        working_date += timedelta(hours=13)
        for row in template:
            for col, value in enumerate(row):
                if col == 3:
                    value = working_date.strftime('%Y/%m/%d %H:%M:%S')
                    print(value)
                sheet.cell(row=row_num+row_offset, column=col+1, value=value)
            random_number = random.randint(3, 6)
            working_date += timedelta(minutes=random_number)
            row_num += 1

    # 將修改後的時間寫回到Excel檔案中
    excel_file.save('2022.xlsx')

if __name__ == '__main__':
    main()
